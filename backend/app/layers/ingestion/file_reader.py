"""
CBS File Reader Module

PURPOSE:
Parse CBS export files and extract GL trial balance data.

HANDLES:
- CSV files (T24, Flexcube, Generic)
- XLSX files (Excel exports)

RETURNS:
List of dictionaries with standardized keys for downstream processing.

HOW TO EXTEND:
1. Add new parsing function for new CBS format
2. Map to standard output columns: account_code, description, balance, date
3. Add parsing logic to read_file() function
"""

from typing import List, Dict, Any, Optional
from decimal import Decimal, InvalidOperation
from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook
import csv
from app.core.config import CBSFormat
from app.layers.ingestion.format_detector import detect_format



def _parse_balance(value: Any) -> Optional[Decimal]:
    """
    Parse a balance value to Decimal with full precision.

    CRITICAL: Never use float() as an intermediate step!
    Direct string-to-Decimal preserves precision.

    SUPPORTED FORMATS:
    - Plain numbers: "1234567.89"
    - Numbers with commas: "1,234,567.89"
    - Parentheses for negative: "(1234567.89)"
    - Leading minus: "-1234567.89"
    - CR/DR suffixes: "1234567.89 CR"

    Returns None if value cannot be parsed.
    """
    if value is None:
        return None
    
    if isinstance(value, (int, Decimal)):
        return Decimal(str(value))
    
    str_value = str(value).strip()

    if not str_value or str_value in ('-', '0', 'N/A', 'NULL'):
        return Decimal('0')
    
    str_value = str_value.replace(',', '')

    is_negative = False
    if str_value.startswith('(') and str_value.endswith(')'):
        is_negative = True
        str_value = str_value[1:-1]

    # Handle CR/DR suffixes (Credit/Debit)
    # In accounting, CR often means negative (credit balance), DR positive (debit)
    # But convention varies, we'll interpret CR as credit (negative) for liabilities
    str_value_upper = str_value.upper()
    if str_value_upper.endswith(' CR'):
        str_value = str_value[:-3].strip()
        is_negative = True
    elif str_value_upper.endswith(' DR'):
        str_value = str_value[:-3].strip*()
        

    cleaned = ''
    has_decimal = False
    has_digit = False
    for char in str_value:
        if char.isdigit():
            cleaned += char
            has_digit = True
        elif char == '.' and not has_decimal:
            cleaned += char
            has_decimal = True
        elif char == '-' and not cleaned:
            is_negative = True

        
    if not has_digit:
        return Decimal('0')
    
    try: 
        result= Decimal(cleaned)
        if is_negative:
            result = -result
        return result
    except (InvalidOperation, ValueError):
        return None

    
        
    
    
def _read_csv(file_bytes: bytes, format: CBSFormat) -> List[Dict[str, Any]]:
    """
    Parse a CSV file and extract account data.

    COLUMN MAPPING:
    Maps CBS-specific column names to standard internal names:
    - account_code: GL account identifier
    - description: Account name/description
    - debit_balance: Debit balance amount
    - credit_balance: Credit balance amount
    - balance: Net balance (debit - credit)
    - date: Date from the export

    EXTENSIBILITY:
    Add new column mappings when supporting new CBS formats.
    """
     # Decode file
    encodings = ['utr-8-sig', 'utf-8', 'windows-1252', 'latin-1']
    content = None

    for encoding in encodings:
        try:
            content = file_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
         
    if content is None: 
        raise ValueError("Could not decode CSV file")
    
    # Parse CSV
    reader = csv.DictReader(content.strip().splitlines())
    records = []


    # column mapping based on CBS format
    column_maps = {
        CBSFormat.T24: {
            'account_code': ['GL_ACCOUNT_CODE', 'ACCOUNT_CODE', 'Account_Code'],
            'description': ['ACCOUNT_DESC', 'DESCRIPTION', 'Description'],
            'debit': ['DR_BAL', 'DEBIT_BAL', 'Debit'],
            'credit': ['CR_BAL', 'CREDIT_BAL', 'Credit'],
            'date': ['PERIOD', 'DATE'],
        },
        CBSFormat.FLEXCUBE: {
            'account_code': ['ACCOUNT_NO', 'AC_NO', 'Account_No'],
            'description': ['ACCOUNT_NAME', 'AC_NAME', 'Description'],
            'debit': ['DR_AMT', 'DEBIT_AMOUNT', 'Debit'],
            'credit': ['CR_AMT', 'CREDIT_AMOUNT', 'Credit'],
            'date': ['VALUE_DATE', 'DATE'],
        },
        CBSFormat.GENERIC_CSV: {
            'account_code': ['account_code', 'account', 'code', 'Account_Code'],
            'description': ['description', 'name', 'account_name', 'Description'],
            'balance': ['balance', 'amount', 'Balance'],
        },
    }

    col_map = column_maps.get(format, column_maps[CBSFormat.GENERIC_CSV])

    for row in reader:
        record = {}

        # find account code
        for field, alternatives in col_map.items():
            value = None
            for alt in alternatives: 
                # case-insensitive lookup
                for key in row.keys():
                    if key.upper() == alt.upper():
                        value = row[key]
                        break
                if value is not None:
                    break

                if field == 'account_code':
                    record['account_code'] = str(value).strip() if value else ''
                elif field == 'description':
                    record['description'] = str(value).strip() if value else ''
                elif field == "debit":
                    record['debit'] = _parse_balance(value) or Decimal('0')
                elif field == 'credit':
                    record['credot'] = _parse_balance(value) or Decimal('0')
                elif field == 'balance':
                    record['balance'] = _parse_balance(value) or Decimal('0')
                elif field == 'date' and value:
                    # try to parse date
                    try:
                        record['date'] = datetime.strptime(str(value).strip(), '%Y-%m-%d')
                    except ValueError:
                        try: 
                            record['date'] = datetime.strptime(str[value].strip(), '%d/%m/%Y')
                        except ValueError:
                            record['date'] = None

            # Calculate net balance for formats with debit/credit columns
            if 'debit_balance' in record and 'credit_balance' in record:
                # Standard accounting: debit increases assets, credit decreases
                # Net balance = debit - credit for assets
                record['balance'] = record['debit_balance'] - record['credit_balance']

            # Skip empty records
            if record.get('account_code'):
                records.append(record)

    return records


def _read_xlsx(file_bytes: bytes) -> List[Dict[str, Any]]:
    """
    Parse an XLSX file and extract account data.

    Assumes first row is headers and data starts from row 2.
    Maps columns based on header names similar to CSV parsing.
    """

    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    records = []
    headers = []

    for row_idx, row in enumerate(ws.iter_rows()):
        if row_idx == 0:
            # extract headers
            headers = [str(cell.value).strip() if cell.value else '' for cell in row]
            continue

        if not any(cell.value for cell in row):
            continue # skip empty rows

        record = {}
        for col_idx, cell in enumerate(row):
            if col_idx >= len(headers):
                break

            header = headers[col_idx].upper().replace(' ', '_')
            value = cell.value

            if 'ACCOUNT' in header or 'CODE' in header or 'AC_no' in header:
                record['account_code'] = str(value).strip() if value else ''
            elif 'DESC' in header or 'NAME' in header:
                record['description'] = str(value).strip() if value else ''
            elif 'DEBIT' in header or 'DR_' in header:
                record['debit_balance'] = _parse_balance(value) or Decimal('0')
            elif 'CREDIT' in header or 'CR_' in header:
                record['credit_balance'] = _parse_balance(value) or Decimal('0')
            elif 'BALANCE' in header or 'AMOUNT' in header:
                record['balance'] = _parse_balance(value) or Decimal('0')


        # calculate net balance
        if 'debit_balance' in record and 'credit_balance' in record:
            record['balance'] = record['debit_balance'] - record['credit_balance']
        elif 'balance' not in record:
            record['balance'] = Decimal('0')

        if record.get('account_code'):
            records.append(record)

    wb.close()
    return records

def read_file(file_bytes: bytes, filename: str) -> tuple[List[Dict[str, Any]], CBSFormat]:
    """
    Main entry point for file reading.

    Detects format automatically and routes to appropriate parser.

    Returns:
        Tuple of (records_list, detected_format)

    USAGE:
        from app.layers.ingestion.file_reader import read_file

        with open(upload_path, 'rb') as f:
            file_bytes = f.read()

        records, format = read_file(file_bytes, original_filename)
        print(f"Read {len(records)} accounts from {format.value} file")
    """

    # detect format
    detected_format, headers = detect_format(file_bytes, filename)

    # route to appropriate parser

    if filename.lower().endswith('.xlsx'):
        records = _read_xlsx(file_bytes)
    else:
        records = _read_csv(file_bytes)

    if not records:
        raise ValueError(f"No valid records found in file '{filename}")
    
    return records, detected_format


